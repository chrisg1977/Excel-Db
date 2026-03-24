from __future__ import annotations

import os
import sys
from dataclasses import dataclass
from decimal import Decimal, ROUND_DOWN
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
import psycopg2


DEFAULT_WORKBOOK = r"C:\Users\User\Desktop\CURRENT BACKUP\PAYROLL.xlsm"
TARGET_PAY_TYPE = "HOURLY_RATE"
TARGET_CURRENCY = "EUR"


def load_local_env() -> None:
    env_path = Path(__file__).resolve().parents[1] / ".env"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if key and key not in os.environ:
            os.environ[key] = value


@dataclass
class RateRow:
    emp_id: int
    effective_from: date
    amount: float
    notes: str


def parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_float(value) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not (number > 0):
        return None
    return float(number)


def truncate_2dp(value: float) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.00"), rounding=ROUND_DOWN))


def extract_rows(path: Path) -> list[RateRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[RateRow] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        marker = ws.cell(2, 13).value
        emp_value = ws.cell(2, 14).value
        if str(marker).strip() != "EMP NO>":
            continue
        try:
            emp_id = int(emp_value)
        except (TypeError, ValueError):
            continue

        for r in range(3, 16):
            amount = parse_float(ws.cell(r, 14).value)
            effective_from = parse_date(ws.cell(r, 15).value)
            if amount is None or effective_from is None:
                continue
            rows.append(
                RateRow(
                    emp_id=emp_id,
                    effective_from=effective_from,
                    amount=amount,
                    notes=f"Imported from {path.name} sheet {sheet_name} row {r}",
                )
            )

    deduped: dict[tuple[int, date], RateRow] = {}
    for row in rows:
        deduped[(row.emp_id, row.effective_from)] = row
    return sorted(deduped.values(), key=lambda item: (item.emp_id, item.effective_from))


def main() -> int:
    load_local_env()
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(DEFAULT_WORKBOOK)
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    conn_str = os.environ.get("DIRECTUS_PG_CONNECTION")
    if not conn_str:
        print("Missing DIRECTUS_PG_CONNECTION", file=sys.stderr)
        return 1

    rows = extract_rows(workbook_path)
    if not rows:
        print("No payroll rates found in workbook", file=sys.stderr)
        return 1

    conn = psycopg2.connect(conn_str)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("SELECT emp_id FROM employees")
                valid_employee_ids = {int(row[0]) for row in cur.fetchall()}
                rows_to_import = [row for row in rows if row.emp_id in valid_employee_ids]
                cur.execute(
                    "DELETE FROM employee_pay_private WHERE pay_type = %s",
                    (TARGET_PAY_TYPE,),
                )
                for row in rows_to_import:
                    cur.execute(
                        """
                        INSERT INTO employee_pay_private (
                          emp_id,
                          pay_type,
                          amount,
                          currency,
                          effective_from,
                          effective_to,
                          notes
                        ) VALUES (%s, %s, %s, %s, %s, NULL, %s)
                        """,
                        (
                            row.emp_id,
                            TARGET_PAY_TYPE,
                            truncate_2dp(row.amount),
                            TARGET_CURRENCY,
                            row.effective_from,
                            row.notes,
                        ),
                    )
        skipped = len(rows) - len(rows_to_import)
        print(f"Imported {len(rows_to_import)} hourly rate row(s) from {workbook_path}; skipped {skipped} row(s) with no live employee match")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
