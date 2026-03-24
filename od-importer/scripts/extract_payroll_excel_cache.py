from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\User\Desktop\CURRENT BACKUP\PAYROLL.xlsm")
TARGET = ROOT / "data" / "payroll_excel_cache.json"


def iso(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def money(value):
    if value is None or value == "":
        return None
    try:
        return round(float(value), 2)
    except Exception:
        return value


def text(value):
    if value is None:
        return ""
    return str(value).strip()


def iban_text(value):
    raw = text(value).replace(" ", "")
    if not raw:
        return ""
    if raw.lower() in {"nan", "0", "#n/a", "null", "undefined"}:
        return ""
    return raw


def extract_block(ws, start_row: int):
    n = lambda c: ws[f"{c}{start_row}"].value
    row = lambda c, offset: ws[f"{c}{start_row + offset}"].value
    month_marker = row("W", 0)
    if not isinstance(month_marker, (date, datetime)):
        return None
    month_key = iso(month_marker)[:7]

    rows = []
    for offset in range(10):
        r = start_row + offset
        vals = [ws[f"{col}{r}"].value for col in ["N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y"]]
        rows.append(
            {
                "row_key": f"r{r}",
                "cells": [iso(v) if isinstance(v, (date, datetime)) else v for v in vals],
            }
        )

    data = {
        "month": month_key,
        "sheet": ws.title,
        "period_from": iso(row("W", 0)),
        "period_to": iso(row("N", 3)),
        "employee_name": text(row("O", 0)),
        "id_card": text(row("P", 0)),
        "designation": text(row("Q", 0)),
        "gross_total": money(row("S", 0)),
        "fs4_status": text(row("T", 0)),
        "employment_type": text(row("U", 0)),
        "ssc_class": text(row("V", 0)),
        "hours_per_week": money(row("Q", 1)),
        "basic_wage": money(row("S", 1)),
        "extra_under_label": text(row("P", 2)),
        "extra_under_hours_text": text(row("R", 2)),
        "extra_under_amount": money(row("S", 2)),
        "ssc_total": money(row("T", 2)),
        "ssc_employer": money(row("U", 2)),
        "ssc_employee": money(row("V", 2)),
        "payer_pe_text": text(row("O", 3)),
        "unpaid_leave_hours_text": text(row("R", 3)),
        "unpaid_leave_amount": money(row("S", 3)),
        "mlf": money(row("U", 3)),
        "bonus": money(row("S", 4)),
        "overtime_tax": money(row("V", 4)),
        "other_amount": money(row("S", 5)),
        "tax_on_gross": money(row("V", 5)),
        "performance_bonus": money(row("P", 6)),
        "supervisor_bonus": money(row("R", 6)),
        "iban": iban_text(row("T", 6)),
        "overtime_hours_text": text(row("Q", 7)),
        "overtime_amount": money(row("S", 7)),
        "net_wage": money(row("U", 7)),
        "leave_taken_to_date": money(row("Q", 8)),
        "transaction_id_text": text(row("Y", 8) or row("T", 8)),
        "weeks_value": money(row("N", 8)),
        "banked_hours_since_last_month": money(row("Q", 9)),
        "leave_balance_end": money(row("S", 9)),
        "clinic_addresses": text(row("T", 9)),
        "rows": rows,
    }
    return data


def main():
    wb = openpyxl.load_workbook(SOURCE, data_only=True, keep_vba=True)
    result = {"source": str(SOURCE), "employees": {}}
    for ws in wb.worksheets:
        try:
            emp_id = ws["O23"].value
        except Exception:
            continue
        if not isinstance(emp_id, int):
            continue

        employee_payload = {"sheet": ws.title, "months": {}}
        for month_index in range(12):
            start_row = 19 + (month_index * 10)
            block = extract_block(ws, start_row)
            if block:
                employee_payload["months"][block["month"]] = block
        result["employees"][str(emp_id)] = employee_payload

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    TARGET.write_text(json.dumps(result, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {TARGET}")


if __name__ == "__main__":
    main()
